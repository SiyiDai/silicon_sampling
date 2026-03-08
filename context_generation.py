from __future__ import annotations

import csv
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

from openpyxl import load_workbook


PERSONA_SOURCE_KEYS = [
    "Q20",
    "Q19_1",
    "Q19_2",
    "Q19_3",
    "Q19_4",
    "Q28",
    "Q29",
    "Q30",
    "Q31",
]


SurveyRow = Dict[str, str]


def _clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _normalize_row(fieldnames: Sequence[str], values: Sequence[object]) -> SurveyRow:
    row: SurveyRow = {}
    for key, value in zip(fieldnames, values):
        row[key] = _clean_text(value)
    return row


def _load_csv_table(csv_file: str) -> Tuple[List[str], List[str], List[SurveyRow]]:
    rows: List[SurveyRow] = []
    with open(csv_file, newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        _ = next(reader)  # Column1..ColumnN
        fieldnames = [_clean_text(cell) for cell in next(reader)]
        labels = [_clean_text(cell) for cell in next(reader)]
        _ = next(reader)  # import metadata

        for values in reader:
            if not values or all(not _clean_text(cell) for cell in values):
                continue
            rows.append(_normalize_row(fieldnames, values))

    return fieldnames, labels, rows


def _load_excel_table(excel_file: str) -> Tuple[List[str], List[str], List[SurveyRow]]:
    workbook = load_workbook(excel_file, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    metadata_rows = list(worksheet.iter_rows(min_row=1, max_row=4, values_only=True))
    if len(metadata_rows) < 4:
        raise ValueError("Expected at least four header rows in the Excel file.")

    fieldnames = [_clean_text(cell) for cell in metadata_rows[1]]
    labels = [_clean_text(cell) for cell in metadata_rows[2]]

    rows: List[SurveyRow] = []
    for values in worksheet.iter_rows(min_row=5, values_only=True):
        if not values or all(not _clean_text(cell) for cell in values):
            continue
        rows.append(_normalize_row(fieldnames, values))

    workbook.close()
    return fieldnames, labels, rows


def read_survey_table(table_file: str) -> Tuple[List[str], List[str], List[SurveyRow]]:
    """
    Read a Qualtrics-style CSV or XLSX survey export.

    Returns:
        fieldnames, labels, rows
    """
    suffix = Path(table_file).suffix.lower()
    if suffix == ".csv":
        return _load_csv_table(table_file)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _load_excel_table(table_file)
    raise ValueError(f"Unsupported file type: {suffix}")


def load_survey_rows(table_file: str) -> Iterable[Tuple[List[str], List[str], SurveyRow]]:
    """
    Backward-compatible row iterator.
    """
    fieldnames, labels, rows = read_survey_table(table_file)
    for row in rows:
        yield fieldnames, labels, row


def build_label_map(fieldnames: Sequence[str], labels: Sequence[str]) -> Dict[str, str]:
    return {
        _clean_text(fieldname): _clean_text(label) or _clean_text(fieldname)
        for fieldname, label in zip(fieldnames, labels)
        if _clean_text(fieldname)
    }


def build_persona_seed_context(
    row: SurveyRow,
    label_map: Dict[str, str],
    persona_keys: Sequence[str] | None = None,
) -> str:
    """
    Build the grounded seed context used to create a persona.
    Only the selected persona_keys are included.
    """
    keys = list(persona_keys or PERSONA_SOURCE_KEYS)
    qa_lines: List[str] = []
    for key in keys:
        answer = _clean_text(row.get(key))
        if not answer:
            continue
        question = label_map.get(key, key)
        qa_lines.append(f"- {key} | {question}: {answer}")

    if not qa_lines:
        return "No persona seed responses available."

    return "\n".join(qa_lines)


def get_target_question_keys(
    fieldnames: Sequence[str],
    excluded_keys: Sequence[str] | None = None,
) -> List[str]:
    excluded = set(excluded_keys or [])
    return [key for key in fieldnames if key.startswith("Q") and key not in excluded]


def build_target_question_map(
    fieldnames: Sequence[str],
    label_map: Dict[str, str],
    excluded_keys: Sequence[str] | None = None,
) -> Dict[str, str]:
    target_keys = get_target_question_keys(fieldnames, excluded_keys=excluded_keys)
    return {key: label_map.get(key, key) for key in target_keys}


def build_target_question_spec(target_questions: Dict[str, str]) -> str:
    lines = []
    for key, label in target_questions.items():
        lines.append(f'- "{key}": {label}')
    return "\n".join(lines)
